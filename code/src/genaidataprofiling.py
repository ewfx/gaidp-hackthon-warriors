import pdfplumber
import requests
import os
import json
import csv
import pandas as pd
from sklearn.ensemble import IsolationForest
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def extract_text_from_pdf(pdf_path):
    """Extracts text from a given PDF file."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text

def clean_json_response(response_text):
    """Extracts valid JSON from LLM response by removing LaTeX-style and code block formatting."""
    match = re.search(r'```json\n(.*?)\n```', response_text, re.DOTALL)
    if match:
        return match.group(1)  # Extract JSON inside code block
    
    match = re.search(r'\\boxed{(.*?)}', response_text, re.DOTALL)
    if match:
        return match.group(1)  # Extract JSON inside \boxed{}
    
    return response_text  # Return original if no formatting detected

def extract_profile_rules(text, model="deepseek/deepseek-r1-zero:free", api_key=None, url="https://openrouter.ai/api/v1/chat/completions"):
    """Uses OpenRouter API to extract structured profiling rules from regulatory text."""
    if api_key is None:
        api_key = os.getenv("DEEPSEEK_API_KEY")  # Ensure the API key is available
    
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    
    prompt = (
        "Extract structured data profiling rules from the following FR Y-14Q regulatory reporting instructions. "
        "Ensure the output is strictly in JSON format as a list of dictionaries, where each dictionary has: "
        "'column' (string, must be unique), 'constraint' (string, e.g., 'non-null', 'greater than', 'allowed values'), and 'value' (number or list). "
        "Ensure each column name is unique. Do not include any text before or after the JSON output. Output only valid JSON."
    )
    
    data = {
        "model": model,
        "messages": [
            {"role": "system", "content": prompt},
            {"role": "user", "content": text}
        ],
        "temperature": 0.2,
        "stream": False
    }
    
    response = requests.post(url, headers=headers, json=data)
    if response.status_code != 200:
        print(f"Error: Received status code {response.status_code}")
        print("Response:", response.text)
        return "[]"  # Return empty JSON list to prevent failure
    
    try:
        result = response.json()
        print("Raw API Response:", result)  # Debugging step
        
        choices = result.get("choices", [])
        if not choices or "message" not in choices[0] or "content" not in choices[0]["message"]:
            print("Error: API response is missing expected fields.")
            return "[]"
        
        rules_text = choices[0]["message"]["content"]
        cleaned_json = clean_json_response(rules_text)  # Remove LaTeX-style formatting
        
        # Ensure the response is valid JSON
        if cleaned_json.strip().startswith("["):
            return cleaned_json
        else:
            print("Error: LLM response is not valid JSON.")
            print("Response:", cleaned_json)
            return "[]"
    except json.JSONDecodeError:
        print("Error: Failed to parse JSON response")
        print("Response:", response.text)
        return "[]"  # Return empty JSON list

def validate_data(data_file, rules, output_file="validation_results.xlsx"):
    """Validates data against structured profiling rules and saves findings along with input data to an Excel file with error highlighting and remediation suggestions."""
    df = pd.read_csv(data_file)
    try:
        rules = json.loads(rules)
    except json.JSONDecodeError:
        print("Error: Invalid JSON format in extracted rules.")
        return
    
    df["Error_Details"] = ""
    
    for rule in rules:
        column, constraint, value = rule.get("column"), rule.get("constraint"), rule.get("value")
        if not column or not constraint:
            continue  # Skip invalid rules
        
        if column in df.columns:
            if constraint == "non-null":
                df.loc[df[column].isnull(), "Error_Details"] += f"{column} is missing. Provide a valid value. "
            elif constraint == "greater than" and pd.api.types.is_numeric_dtype(df[column]):
                df.loc[df[column] <= float(value), "Error_Details"] += f"{column} must be greater than {value}. "
            elif constraint == "allowed values" and isinstance(value, list):
                df.loc[~df[column].isin(value), "Error_Details"] += f"{column} must be one of {value}. "
        else:
            df["Error_Details"] += f"{column} is missing in the dataset. "
    
    # Save to Excel file
    df.to_excel(output_file, index=False)
    
    # Apply red highlight to error records
    wb = load_workbook(output_file)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    error_col_idx = df.columns.get_loc("Error_Details") + 1  # Get correct column index
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=df.shape[1]):
        for cell in row:
            if cell.column == error_col_idx and cell.value:
                cell.fill = red_fill
    
    wb.save(output_file)
    print(f"Validation results saved to {output_file} with error records highlighted in red and error details included.")

# Example usage
pdf_path = "/content/sample_data/FR_Y_14Q_Instructions_wholesale.pdf"  # Replace with actual file path
text = extract_text_from_pdf(pdf_path)
profiling_rules = extract_profile_rules(text, api_key="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")

print("Extracted Profiling Rules:")
print(profiling_rules)

# Validate data and save findings alongside input data with red-highlighted errors
validate_data("/content/sample_data/FR_Y_14Q_Input_Data_Latest.csv", profiling_rules)
